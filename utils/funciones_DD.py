import openpyxl
from openpyxl import Workbook


class data_Driven ():
    
    def contar_Filas(self,path,sheetname):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook[sheetname]
        return sheet.max_row
    
    def contar_Columnas(self,path,sheetmname):
        woorkbok=openpyxl.load_workbook(path)
        sheet=woorkbok[sheetmname]
        return sheet.max_column
    
    def leer_Datos(self,path,sheetmname,num_fila,num_columna):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook[sheetmname]
        return sheet.cell(row=num_fila,column=num_columna).value
    
    def insertar_Datos(self,path,sheetname,num_fila,num_columna,valores):
        workbook=openpyxl.load_workbook(path)
        sheet=workbook[sheetname]
        sheet.cell(row=num_fila,column=num_columna).value=valores
        workbook.save(path)
    
        
